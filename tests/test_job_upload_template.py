"""Module to test job upload template configs and generation"""

import os
import unittest
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from aind_data_transfer_service.configs.job_upload_template import (
    JobUploadTemplate,
)

TEST_DIRECTORY = Path(os.path.dirname(os.path.realpath(__file__)))
SAMPLE_JOB_TEMPLATE = TEST_DIRECTORY / "resources" / "job_upload_template.xlsx"


class TestJobUploadTemplate(unittest.TestCase):
    """Tests job upload template class"""

    @staticmethod
    def _read_xl_helper(source, return_validators=False):
        """Helper function to read xlsx contents and validators"""
        lines = []
        workbook = load_workbook(source, read_only=(not return_validators))
        worksheet = workbook.active
        for row in worksheet.iter_rows(values_only=True):
            lines.append(row) if any(row) else None
        if return_validators:
            validators = []
            for dv in worksheet.data_validations.dataValidation:
                validator = {
                    "name": dv.promptTitle,
                    "type": dv.type,
                    "ranges": str(dv.cells).split(" "),
                }
                if dv.type == "list":
                    validator["options"] = dv.formula1.strip('"').split(",")
                validators.append(validator)
            result = (lines, validators)
        else:
            result = lines
        workbook.close()
        return result

    @classmethod
    def setUpClass(cls):
        """Set up test class"""
        expected_lines = cls._read_xl_helper(SAMPLE_JOB_TEMPLATE)
        (template_lines, template_validators) = cls._read_xl_helper(
            JobUploadTemplate.create_excel_sheet_filestream(), True
        )
        cls.expected_lines = expected_lines
        cls.template_lines = template_lines
        cls.template_validators = template_validators
        return cls

    def test_create_job_template(self):
        """Tests that xlsx job template is created with
        correct contents and validators"""
        expected_lines = self.expected_lines
        self.assertEqual(expected_lines, self.template_lines)
        for validator in self.template_validators:
            validator["column_indexes"] = []
            for r in validator["ranges"]:
                rb = (col, *_) = range_boundaries(r)
                self.assertTupleEqual(
                    (col, 2, col, JobUploadTemplate._NUM_TEMPLATE_ROWS), rb
                )
                validator["column_indexes"].append(col - 1)
            del validator["ranges"]
        self.assertCountEqual(
            JobUploadTemplate._get_validators(), self.template_validators
        )


if __name__ == "__main__":
    unittest.main()
