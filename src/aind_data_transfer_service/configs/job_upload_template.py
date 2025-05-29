"""Module to configure and create xlsx job upload template"""

import datetime
from io import BytesIO
from typing import Any, ClassVar, Dict, List

from aind_data_schema_models.modalities import Modality
from aind_data_schema_models.platforms import Platform
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel


class JobUploadTemplate(BaseModel):
    """Class to configure and create xlsx job upload template"""

    FILE_NAME: ClassVar[str] = "job_upload_template.xlsx"
    _NUM_TEMPLATE_ROWS: ClassVar[int] = 20
    _XLSX_DATETIME_FORMAT: ClassVar[str] = "YYYY-MM-DDTHH:mm:ss"
    _HEADERS: ClassVar[List[str]] = [
        "job_type",
        "project_name",
        "platform",
        "acq_datetime",
        "subject_id",
        "metadata_dir",
        "modality0",
        "modality0.input_source",
        "modality1",
        "modality1.input_source",
    ]
    _SAMPLE_JOBS: ClassVar[List[List[Any]]] = [
        [
            "default",
            "Behavior Platform",
            Platform.BEHAVIOR.abbreviation,
            datetime.datetime(2023, 10, 4, 4, 0, 0),
            "123456",
            "/allen/aind/stage/fake/metadata_dir",
            Modality.BEHAVIOR_VIDEOS.abbreviation,
            "/allen/aind/stage/fake/dir",
            Modality.BEHAVIOR.abbreviation,
            "/allen/aind/stage/fake/dir",
        ],
        [
            "default",
            "Ophys Platform - SLAP2",
            Platform.SMARTSPIM.abbreviation,
            datetime.datetime(2023, 3, 4, 16, 30, 0),
            "654321",
            "/allen/aind/stage/fake/Config",
            Modality.SPIM.abbreviation,
            "/allen/aind/stage/fake/dir",
        ],
        [
            "default",
            "Ephys Platform",
            Platform.ECEPHYS.abbreviation,
            datetime.datetime(2023, 1, 30, 19, 1, 0),
            "654321",
            None,
            Modality.ECEPHYS.abbreviation,
            "/allen/aind/stage/fake/dir",
            Modality.BEHAVIOR_VIDEOS.abbreviation,
            "/allen/aind/stage/fake/dir",
        ],
    ]

    @classmethod
    def _get_validators(cls) -> List[Dict[str, Any]]:
        """
        Returns
        -------
        List[Dict[str, Any]]
          A list of validators for fields that require validation.

        """
        return [
            {
                "name": "platform",
                "type": "list",
                "options": list(Platform.abbreviation_map.keys()),
                "column_indexes": [cls._HEADERS.index("platform")],
            },
            {
                "name": "modality",
                "type": "list",
                "options": list(Modality.abbreviation_map.keys()),
                "column_indexes": [
                    cls._HEADERS.index("modality0"),
                    cls._HEADERS.index("modality1"),
                ],
            },
            {
                "name": "datetime",
                "type": "date",
                "column_indexes": [cls._HEADERS.index("acq_datetime")],
            },
        ]

    @classmethod
    def create_excel_sheet_filestream(cls) -> BytesIO:
        """Create job template as xlsx filestream"""
        xl_io = BytesIO()
        workbook = Workbook()
        workbook.iso_dates = True
        worksheet = workbook.active
        worksheet.append(cls._HEADERS)
        for job in cls._SAMPLE_JOBS:
            worksheet.append(job)
        # data validators
        for validator in cls._get_validators():
            dv_type = validator["type"]
            dv_name = validator["name"]
            dv_params = {
                "type": dv_type,
                "promptTitle": dv_name,
                "error": f"Invalid {dv_name}.",
                "allow_blank": True,
                "showErrorMessage": True,
                "showInputMessage": True,
            }
            if dv_type == "list":
                dv_params["formula1"] = f'"{(",").join(validator["options"])}"'
                dv_params["prompt"] = f"Select a {dv_name} from the dropdown"
            elif dv_type == "date":
                dv_params["prompt"] = "Provide a {} using {}".format(
                    dv_name, cls._XLSX_DATETIME_FORMAT
                )
            dv = DataValidation(**dv_params)
            for i in validator["column_indexes"]:
                col = get_column_letter(i + 1)
                col_range = f"{col}2:{col}{cls._NUM_TEMPLATE_ROWS}"
                dv.add(col_range)
                if dv_type != "date":
                    continue
                for (cell,) in worksheet[col_range]:
                    cell.number_format = cls._XLSX_DATETIME_FORMAT
            worksheet.add_data_validation(dv)
        # formatting
        bold = Font(bold=True)
        for cell in worksheet[1]:
            cell.font = bold
            worksheet.column_dimensions[cell.column_letter].auto_size = True
        # save file
        workbook.save(xl_io)
        workbook.close()
        return xl_io
