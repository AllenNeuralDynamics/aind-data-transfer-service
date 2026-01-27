"""Starts and Runs Starlette Service"""

import csv
import io
import json
import logging
import os
import re
from asyncio import gather
from typing import Any, List, Optional, Union

import boto3
from authlib.integrations.starlette_client import OAuth
from botocore.exceptions import ClientError
from fastapi import Request
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from httpx import AsyncClient
from openpyxl import load_workbook
from pydantic import ValidationError
from starlette.applications import Starlette
from starlette.config import Config
from starlette.middleware.sessions import SessionMiddleware
from starlette.responses import RedirectResponse
from starlette.routing import Route

from aind_data_transfer_service import (
    __version__ as aind_data_transfer_service_version,
)
from aind_data_transfer_service.configs.csv_handler import map_csv_row_to_job
from aind_data_transfer_service.configs.job_upload_template import (
    JobUploadTemplate,
)
from aind_data_transfer_service.log_handler import log_submit_job_request
from aind_data_transfer_service.models.core import (
    SubmitJobRequestV2,
    validation_context,
)
from aind_data_transfer_service.models.internal import (
    AirflowDagRunsRequestParameters,
    AirflowDagRunsResponse,
    AirflowTaskInstanceLogsRequestParameters,
    AirflowTaskInstancesRequestParameters,
    AirflowTaskInstancesResponse,
    JobParamInfo,
    JobStatus,
    JobTasks,
)

template_directory = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "templates")
)
templates = Jinja2Templates(directory=template_directory)

# TODO: Add server configs model
# AIND_METADATA_SERVICE_PROJECT_NAMES_URL
# AIND_AIRFLOW_SERVICE_URL
# AIND_AIRFLOW_SERVICE_JOBS_URL
# AIND_AIRFLOW_SERVICE_PASSWORD
# AIND_AIRFLOW_SERVICE_USER
# ENV_NAME
# LOG_LEVEL

project_names_url = os.getenv("AIND_METADATA_SERVICE_PROJECT_NAMES_URL")


async def validate_csv(request: Request):
    """Validate a csv or xlsx file. Return parsed contents as json."""
    logging.info("Received request to validate csv")
    async with request.form() as form:
        basic_jobs = []
        errors = []
        if not form["file"].filename.endswith((".csv", ".xlsx")):
            errors.append("Invalid input file type")
        else:
            content = await form["file"].read()
            if form["file"].filename.endswith(".csv"):
                # A few csv files created from excel have extra unicode
                # byte chars. Adding "utf-8-sig" should remove them.
                data = content.decode("utf-8-sig")
            else:
                xlsx_book = load_workbook(io.BytesIO(content), read_only=True)
                xlsx_sheet = xlsx_book.active
                csv_io = io.StringIO()
                csv_writer = csv.writer(csv_io)
                for r in xlsx_sheet.iter_rows(values_only=True):
                    if any(r):
                        csv_writer.writerow(r)
                xlsx_book.close()
                data = csv_io.getvalue()
            csv_reader = csv.DictReader(io.StringIO(data))
            params = AirflowDagRunsRequestParameters(
                dag_ids=["transform_and_upload_v2", "run_list_of_jobs"],
                states=["running", "queued"],
            )
            _, current_jobs = await get_airflow_jobs(
                params=params, get_confs=True
            )
            context = {
                "job_types": get_job_types("v2"),
                "project_names": await get_project_names(),
                "current_jobs": current_jobs,
            }
            for row in csv_reader:
                if not any(row.values()):
                    continue
                try:
                    with validation_context(context):
                        job = map_csv_row_to_job(row=row)
                    # Construct hpc job setting most of the vars from the env
                    basic_jobs.append(
                        json.loads(
                            job.model_dump_json(
                                round_trip=True,
                                exclude_none=True,
                                warnings=False,
                            )
                        )
                    )
                except ValidationError as e:
                    errors.append(e.json())
                except Exception as e:
                    errors.append(f"{str(e.args)}")
        message = "There were errors" if len(errors) > 0 else "Valid Data"
        status_code = 406 if len(errors) > 0 else 200
        content = {
            "message": message,
            "data": {"jobs": basic_jobs, "errors": errors},
        }
        return JSONResponse(
            content=content,
            status_code=status_code,
        )


async def get_project_names() -> List[str]:
    """Get a list of project_names"""
    # TODO: Cache response for 5 minutes
    async with AsyncClient() as async_client:
        response = await async_client.get(project_names_url)
        response.raise_for_status()
        project_names = response.json()["data"]
    return project_names


def set_oauth() -> OAuth:
    """Set up OAuth for the service"""
    secrets_client = boto3.client("secretsmanager")
    secret_response = secrets_client.get_secret_value(
        SecretId=os.getenv("AIND_SSO_SECRET_NAME")
    )
    secret_value = json.loads(secret_response["SecretString"])
    for secrets in secret_value:
        os.environ[secrets] = secret_value[secrets]
    config = Config()
    oauth = OAuth(config)
    oauth.register(
        name="azure",
        client_id=config("CLIENT_ID"),
        client_secret=config("CLIENT_SECRET"),
        server_metadata_url=config("AUTHORITY"),
        client_kwargs={"scope": "openid email profile"},
    )
    return oauth


def get_job_types(version: Optional[str] = None) -> List[str]:
    """Get a list of job_types"""
    params = get_parameter_infos(version)
    job_types = list(set([p.job_type for p in params]))
    return job_types


def get_parameter_infos(version: Optional[str] = None) -> List[JobParamInfo]:
    """Get a list of job_type parameters"""
    ssm_client = boto3.client("ssm")
    paginator = ssm_client.get_paginator("describe_parameters")
    params_iterator = paginator.paginate(
        ParameterFilters=[
            {
                "Key": "Path",
                "Option": "Recursive",
                "Values": [JobParamInfo.get_parameter_prefix(version)],
            }
        ]
    )
    params = []
    param_regex = JobParamInfo.get_parameter_regex(version)
    for page in params_iterator:
        for param in page["Parameters"]:
            if match := re.match(param_regex, param.get("Name")):
                param_info = JobParamInfo.from_aws_describe_parameter(
                    parameter=param,
                    job_type=match.group("job_type"),
                    task_id=match.group("task_id"),
                    modality=match.group("modality"),
                    version=version,
                )
                params.append(param_info)
            else:
                logging.info(f"Ignoring {param.get('Name')}")
    return params


def get_parameter_value(param_name: str) -> dict:
    """Get a parameter value from AWS param store based on parameter name"""
    ssm_client = boto3.client("ssm")
    param_response = ssm_client.get_parameter(
        Name=param_name, WithDecryption=True
    )
    param_value = json.loads(param_response["Parameter"]["Value"])
    return param_value


def put_parameter_value(param_name: str, param_value: dict) -> Any:
    """Set a parameter value in AWS param store based on parameter name"""
    param_value_str = json.dumps(param_value)
    ssm_client = boto3.client("ssm")
    result = ssm_client.put_parameter(
        Name=param_name,
        Value=param_value_str,
        Type="String",
        Overwrite=True,
    )
    return result


async def get_airflow_jobs(
    params: AirflowDagRunsRequestParameters, get_confs: bool = False
) -> tuple[int, Union[List[JobStatus], List[dict]]]:
    """Get Airflow jobs using input query params. If get_confs is true,
    only the job conf dictionaries are returned."""

    async def fetch_jobs(
        client: AsyncClient, url: str, request_body: dict
    ) -> tuple[int, Union[List[JobStatus], List[dict]]]:
        """Helper method to fetch jobs using httpx async client"""
        response = await client.post(url, json=request_body)
        response.raise_for_status()
        response_jobs = response.json()
        dag_runs = AirflowDagRunsResponse.model_validate_json(
            json.dumps(response_jobs)
        )
        if get_confs:
            jobs_list = [d.conf for d in dag_runs.dag_runs if d.conf]
        else:
            jobs_list = [
                JobStatus.from_airflow_dag_run(d) for d in dag_runs.dag_runs
            ]
        total_entries = dag_runs.total_entries
        return (total_entries, jobs_list)

    airflow_url = os.getenv("AIND_AIRFLOW_SERVICE_JOBS_URL", "").strip("/")
    airflow_url = f"{airflow_url}/~/dagRuns/list"
    params_dict = json.loads(params.model_dump_json(exclude_none=True))
    # Send request to Airflow to ListDagRuns
    async with AsyncClient(
        auth=(
            os.getenv("AIND_AIRFLOW_SERVICE_USER"),
            os.getenv("AIND_AIRFLOW_SERVICE_PASSWORD"),
        )
    ) as async_client:
        # Fetch initial jobs
        (total_entries, jobs_list) = await fetch_jobs(
            client=async_client,
            url=airflow_url,
            request_body=params_dict,
        )
        # Fetch remaining jobs concurrently
        tasks = list()
        offset = params_dict["page_offset"] + params_dict["page_limit"]
        while offset < total_entries:
            batch_params = {**params_dict, "page_offset": offset}
            tasks.append(
                fetch_jobs(
                    client=async_client,
                    url=airflow_url,
                    request_body=batch_params,
                )
            )
            offset += params_dict["page_limit"]
        batches = await gather(*tasks)
        for _, jobs_batch in batches:
            jobs_list.extend(jobs_batch)
    return (total_entries, jobs_list)


async def validate_json_v2(request: Request):
    """Validate raw json against data transfer models. Returns validated
    json or errors if request is invalid."""
    logging.info("Received request to validate json v2")
    content = await request.json()
    try:
        log_submit_job_request(content=content)
        params = AirflowDagRunsRequestParameters(
            dag_ids=["transform_and_upload_v2", "run_list_of_jobs"],
            states=["running", "queued"],
        )
        _, current_jobs = await get_airflow_jobs(params=params, get_confs=True)
        context = {
            "job_types": get_job_types("v2"),
            "project_names": await get_project_names(),
            "current_jobs": current_jobs,
        }
        with validation_context(context):
            validated_model = SubmitJobRequestV2.model_validate_json(
                json.dumps(content)
            )
        validated_content = json.loads(
            validated_model.model_dump_json(warnings=False, exclude_none=True)
        )
        logging.info("Valid model detected")
        return JSONResponse(
            status_code=200,
            content={
                "message": "Valid model",
                "data": {
                    "version": aind_data_transfer_service_version,
                    "model_json": content,
                    "validated_model_json": validated_content,
                },
            },
        )
    except ValidationError as e:
        logging.warning(f"There were validation errors processing {content}")
        return JSONResponse(
            status_code=406,
            content={
                "message": "There were validation errors",
                "data": {
                    "version": aind_data_transfer_service_version,
                    "model_json": content,
                    "errors": e.json(),
                },
            },
        )
    except Exception as e:
        logging.exception(e, exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "message": "There was an internal server error",
                "data": {
                    "version": aind_data_transfer_service_version,
                    "model_json": content,
                    "errors": str(e.args),
                },
            },
        )


async def submit_jobs_v2(request: Request):
    """Post SubmitJobRequestV2 raw json to hpc server to process."""
    logging.info("Received request to submit jobs v2")
    content = await request.json()
    try:

        params = AirflowDagRunsRequestParameters(
            dag_ids=["transform_and_upload_v2", "run_list_of_jobs"],
            states=["running", "queued"],
        )
        _, current_jobs = await get_airflow_jobs(params=params, get_confs=True)
        context = {
            "job_types": get_job_types("v2"),
            "project_names": await get_project_names(),
            "current_jobs": current_jobs,
        }
        with validation_context(context):
            model = SubmitJobRequestV2.model_validate_json(json.dumps(content))
        full_content = json.loads(
            model.model_dump_json(warnings=False, exclude_none=True)
        )
        logging.info(
            f"Valid request detected. Sending list of jobs. "
            f"dag_id: {model.dag_id}"
        )
        total_jobs = len(model.upload_jobs)
        for job_index, job in enumerate(model.upload_jobs, 1):
            logging.info(
                f"{job.job_type}, {job.s3_prefix} sending to airflow. "
                f"{job_index} of {total_jobs}."
            )

        async with AsyncClient(
            auth=(
                os.getenv("AIND_AIRFLOW_SERVICE_USER"),
                os.getenv("AIND_AIRFLOW_SERVICE_PASSWORD"),
            )
        ) as async_client:
            response = await async_client.post(
                url=os.getenv("AIND_AIRFLOW_SERVICE_URL"),
                json={"conf": full_content},
            )
            status_code = response.status_code
            response_json = response.json()
        return JSONResponse(
            status_code=status_code,
            content={
                "message": "Submitted request to airflow",
                "data": {"responses": [response_json], "errors": []},
            },
        )
    except ValidationError as e:
        logging.warning(f"There were validation errors processing {content}")
        return JSONResponse(
            status_code=406,
            content={
                "message": "There were validation errors",
                "data": {"responses": [], "errors": e.json()},
            },
        )
    except Exception as e:
        logging.exception(e, exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "message": "There was an internal server error",
                "data": {"responses": [], "errors": str(e.args)},
            },
        )


async def get_job_status_list(request: Request):
    """Get status of jobs using input query params."""

    try:
        params = AirflowDagRunsRequestParameters.from_query_params(
            request.query_params
        )
        params_dict = json.loads(params.model_dump_json(exclude_none=True))
        total_entries, job_status_list = await get_airflow_jobs(params=params)
        status_code = 200
        message = "Retrieved job status list from airflow"
        data = {
            "params": params_dict,
            "total_entries": total_entries,
            "job_status_list": [
                json.loads(j.model_dump_json()) for j in job_status_list
            ],
        }
    except ValidationError as e:
        logging.warning(
            f"There was a validation error process job_status list: {e}"
        )
        status_code = 406
        message = "Error validating request parameters"
        data = {"errors": json.loads(e.json())}
    except Exception as e:
        logging.exception(e, exc_info=True)
        status_code = 500
        message = "Unable to retrieve job status list from airflow"
        data = {"errors": [f"{e.__class__.__name__}{e.args}"]}
    return JSONResponse(
        status_code=status_code,
        content={
            "message": message,
            "data": data,
        },
    )


async def get_tasks_list(request: Request):
    """Get list of tasks instances given dag id and job id."""
    try:
        url = os.getenv("AIND_AIRFLOW_SERVICE_JOBS_URL", "").strip("/")
        params = AirflowTaskInstancesRequestParameters.from_query_params(
            request.query_params
        )
        params_dict = json.loads(params.model_dump_json())
        async with AsyncClient(
            auth=(
                os.getenv("AIND_AIRFLOW_SERVICE_USER"),
                os.getenv("AIND_AIRFLOW_SERVICE_PASSWORD"),
            )
        ) as async_client:
            response_tasks = await async_client.get(
                url=(
                    f"{url}/{params.dag_id}/dagRuns/{params.dag_run_id}/"
                    "taskInstances"
                ),
            )
            status_code = response_tasks.status_code
            response_json = response_tasks.json()
        if status_code == 200:
            task_instances = AirflowTaskInstancesResponse.model_validate_json(
                json.dumps(response_json)
            )
            job_tasks_list = sorted(
                [
                    JobTasks.from_airflow_task_instance(t)
                    for t in task_instances.task_instances
                ],
                key=lambda t: (t.priority_weight, t.map_index),
            )
            message = "Retrieved job tasks list from airflow"
            data = {
                "params": params_dict,
                "total_entries": task_instances.total_entries,
                "job_tasks_list": [
                    json.loads(t.model_dump_json()) for t in job_tasks_list
                ],
            }
        else:
            message = "Error retrieving job tasks list from airflow"
            data = {
                "params": params_dict,
                "errors": [response_json],
            }
    except ValidationError as e:
        logging.warning(f"There was a validation error process task_list: {e}")
        status_code = 406
        message = "Error validating request parameters"
        data = {"errors": json.loads(e.json())}
    except Exception as e:
        logging.exception(e, exc_info=True)
        status_code = 500
        message = "Unable to retrieve job tasks list from airflow"
        data = {"errors": [f"{e.__class__.__name__}{e.args}"]}
    return JSONResponse(
        status_code=status_code,
        content={
            "message": message,
            "data": data,
        },
    )


async def get_task_logs(request: Request):
    """Get task logs given dag id, job id, task id, and task try number."""
    try:
        url = os.getenv("AIND_AIRFLOW_SERVICE_JOBS_URL", "").strip("/")
        params = AirflowTaskInstanceLogsRequestParameters.from_query_params(
            request.query_params
        )
        params_dict = json.loads(params.model_dump_json())
        params_full = dict(params)
        async with AsyncClient(
            auth=(
                os.getenv("AIND_AIRFLOW_SERVICE_USER"),
                os.getenv("AIND_AIRFLOW_SERVICE_PASSWORD"),
            )
        ) as async_client:
            response_logs = await async_client.get(
                url=(
                    f"{url}/{params.dag_id}/dagRuns/{params.dag_run_id}"
                    f"/taskInstances/{params.task_id}/logs/{params.try_number}"
                ),
                params=params_dict,
            )
            status_code = response_logs.status_code
            if status_code == 200:
                message = "Retrieved task logs from airflow"
                data = {"params": params_full, "logs": response_logs.text}
            else:
                message = "Error retrieving task logs from airflow"
                data = {
                    "params": params_full,
                    "errors": [response_logs.json()],
                }
    except ValidationError as e:
        logging.warning(f"Error validating request parameters: {e}")
        status_code = 406
        message = "Error validating request parameters"
        data = {"errors": json.loads(e.json())}
    except Exception as e:
        logging.exception(e, exc_info=True)
        status_code = 500
        message = "Unable to retrieve task logs from airflow"
        data = {"errors": [f"{e.__class__.__name__}{e.args}"]}
    return JSONResponse(
        status_code=status_code,
        content={
            "message": message,
            "data": data,
        },
    )


async def index(request: Request):
    """GET|POST /: form handler"""
    user = request.session.get("user")
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context=(
            {
                "user_signed_in": user is not None,
                "project_names_url": project_names_url,
            }
        ),
    )


async def job_tasks_table(request: Request):
    """Get Job Tasks table given a job id"""
    response_tasks = await get_tasks_list(request)
    response_tasks_json = json.loads(response_tasks.body)
    data = response_tasks_json.get("data")
    return templates.TemplateResponse(
        request=request,
        name="job_tasks_table.html",
        context=(
            {
                "status_code": response_tasks.status_code,
                "message": response_tasks_json.get("message"),
                "errors": data.get("errors", []),
                "total_entries": data.get("total_entries", 0),
                "job_tasks_list": data.get("job_tasks_list", []),
            }
        ),
    )


async def task_logs(request: Request):
    """Get task logs given a job id, task id, and task try number."""
    response_tasks = await get_task_logs(request)
    response_tasks_json = json.loads(response_tasks.body)
    data = response_tasks_json.get("data")
    return templates.TemplateResponse(
        request=request,
        name="task_logs.html",
        context=(
            {
                "status_code": response_tasks.status_code,
                "message": response_tasks_json.get("message"),
                "errors": data.get("errors", []),
                "logs": data.get("logs"),
            }
        ),
    )


async def jobs(request: Request):
    """Get Job Status page with pagination"""
    user = request.session.get("user")
    dag_ids = AirflowDagRunsRequestParameters.model_fields["dag_ids"].default
    return templates.TemplateResponse(
        request=request,
        name="job_status.html",
        context=(
            {
                "user_signed_in": user is not None,
                "project_names_url": project_names_url,
                "dag_ids": dag_ids,
            }
        ),
    )


async def job_params(request: Request):
    """Get Job Parameters page"""
    user = request.session.get("user")
    return templates.TemplateResponse(
        request=request,
        name="job_params.html",
        context=(
            {
                "user_signed_in": user is not None,
                "project_names_url": os.getenv(
                    "AIND_METADATA_SERVICE_PROJECT_NAMES_URL"
                ),
                "default_version": "v2",
                "modalities": JobParamInfo._MODALITIES_LIST,
                "modality_tasks": JobParamInfo._MODALITY_TASKS,
            }
        ),
    )


async def download_job_template(_: Request):
    """Get job template as xlsx filestream for download"""

    try:
        xl_io = JobUploadTemplate.create_excel_sheet_filestream()
        return StreamingResponse(
            io.BytesIO(xl_io.getvalue()),
            media_type=(
                "application/"
                "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": (
                    f"attachment; filename={JobUploadTemplate.FILE_NAME}"
                )
            },
            status_code=200,
        )
    except Exception as e:
        logging.exception(e, exc_info=True)
        return JSONResponse(
            content={
                "message": "Error creating job template",
                "data": {"error": f"{e.__class__.__name__}{e.args}"},
            },
            status_code=500,
        )


def list_parameters_v2(_: Request):
    """List v2 job type parameters"""
    params = get_parameter_infos("v2")
    return JSONResponse(
        content={
            "message": "Retrieved job parameters",
            "data": [p.model_dump(mode="json") for p in params],
        },
        status_code=200,
    )


def get_parameter_v2(request: Request):
    """Get v2 parameter from AWS param store based on job_type and task_id"""
    # path params are auto validated
    job_type = request.path_params.get("job_type")
    task_id = request.path_params.get("task_id")
    modality = request.path_params.get("modality")
    param_name = JobParamInfo.get_parameter_name(
        job_type=job_type, task_id=task_id, modality=modality, version="v2"
    )
    try:
        param_value = get_parameter_value(param_name)
        return JSONResponse(
            content={
                "message": f"Retrieved parameter for {param_name}",
                "data": param_value,
            },
            status_code=200,
        )
    except ClientError as e:
        logging.exception(e, exc_info=True)
        return JSONResponse(
            content={
                "message": f"Error retrieving parameter {param_name}",
                "data": {"error": f"{e.__class__.__name__}{e.args}"},
            },
            status_code=500,
        )


async def put_parameter(request: Request):
    """Set parameter in AWS param store based on job_type and task_id"""
    # User must be signed in
    user = request.session.get("user")
    if not user:
        return JSONResponse(
            content={
                "message": "User not authenticated",
                "data": {"error": "User not authenticated"},
            },
            status_code=401,
        )
    try:
        # path params
        param_info = JobParamInfo(
            name=None,
            last_modified=None,
            job_type=request.path_params.get("job_type"),
            task_id=request.path_params.get("task_id"),
            modality=request.path_params.get("modality"),
            version=request.path_params.get("version"),
        )
        param_name = JobParamInfo.get_parameter_name(
            job_type=param_info.job_type,
            task_id=param_info.task_id,
            modality=param_info.modality,
            version=param_info.version,
        )
        # update param store
        logging.info(
            f"Received request from {user} to set parameter {param_name}"
        )
        param_value = await request.json()
        logging.info(f"Setting parameter {param_name} to {param_value}")
        result = put_parameter_value(
            param_name=param_name, param_value=param_value
        )
        logging.info(result)
        return JSONResponse(
            content={
                "message": f"Set parameter for {param_name}",
                "data": param_value,
            },
            status_code=200,
        )
    except ValidationError as error:
        return JSONResponse(
            content={
                "message": "Invalid parameter",
                "data": {"errors": json.loads(error.json())},
            },
            status_code=400,
        )
    except Exception as e:
        logging.exception(e, exc_info=True)
        return JSONResponse(
            content={
                "message": f"Error setting parameter {param_name}",
                "data": {"error": f"{e.__class__.__name__}{e.args}"},
            },
            status_code=500,
        )


async def admin(request: Request):
    """Get admin page if authenticated, else redirect to login."""
    user = request.session.get("user")
    if user:
        return templates.TemplateResponse(
            request=request,
            name="admin.html",
            context=(
                {
                    "user_signed_in": user is not None,
                    "user_name": user.get("name", "unknown"),
                    "user_email": user.get("email", "unknown"),
                    "project_names_url": project_names_url,
                }
            ),
        )
    return RedirectResponse(url="/login")


async def login(request: Request):
    """Redirect to Azure login page"""
    if os.getenv("ENV_NAME") == "local":
        request.session["user"] = {"name": "local user"}
        return RedirectResponse(url="/admin")
    oauth = set_oauth()
    redirect_uri = request.url_for("auth")
    response = await oauth.azure.authorize_redirect(request, redirect_uri)
    return response


async def logout(request: Request):
    """Logout user and clear session"""
    request.session.pop("user", None)
    return RedirectResponse(url="/")


async def auth(request: Request):
    """Authenticate user and store user info in session"""
    oauth = set_oauth()
    try:
        token = await oauth.azure.authorize_access_token(request)
        user = token.get("userinfo")
        if not user:
            raise ValueError("User info not found in access token.")
        request.session["user"] = dict(user)
    except Exception as error:
        return JSONResponse(
            content={
                "message": "Error Logging In",
                "data": {"error": f"{error.__class__.__name__}{error.args}"},
            },
            status_code=500,
        )
    return RedirectResponse(url="/admin")


async def cancel_job(request: Request):
    """Cancel a running job."""
    request_json = await request.json()
    logging.info(f"Received request to cancel job: {request_json}")
    try:
        dag_run_id = request_json["dag_run_id"]
        dag_id = request_json["dag_id"]
        s3_prefix = request_json["s3_prefix"]
        partition = request_json.get("partition", "aind")
        airflow_url = os.getenv("AIND_AIRFLOW_SERVICE_JOBS_URL", "").strip("/")
        cancel_slurm_jobs_DAG_ID = os.getenv(
            "AIND_AIRFLOW_SERVICE_CANCEL_JOBS_DAG_ID", "cancel_slurm_jobs"
        )
        async with AsyncClient(
            auth=(
                os.getenv("AIND_AIRFLOW_SERVICE_USER"),
                os.getenv("AIND_AIRFLOW_SERVICE_PASSWORD"),
            )
        ) as async_client:
            cancel_dag_url = f"{airflow_url}/{dag_id}/dagRuns/{dag_run_id}"
            cancel_dag_response = await async_client.patch(
                url=cancel_dag_url, json={"state": "failed"}
            )
            cancel_dag_response.raise_for_status()
            cancel_slurm_jobs_url = (
                f"{airflow_url}/{cancel_slurm_jobs_DAG_ID}/dagRuns"
            )
            cancel_slurm_jobs_response = await async_client.post(
                url=cancel_slurm_jobs_url,
                json={
                    "conf": {
                        "dag_run_id": dag_run_id,
                        "s3_prefix": s3_prefix,
                        "partition": partition,
                    }
                },
            )
            cancel_slurm_jobs_response.raise_for_status()
            return JSONResponse(
                content={
                    "message": "Success",
                    "data": dict(),
                },
                status_code=200,
            )
    except Exception as e:
        logging.exception(e, exc_info=True)
        return JSONResponse(
            content={
                "message": "Error canceling job.",
                "data": {"error": str(e)},
            },
            status_code=500,
        )


routes = [
    Route("/", endpoint=index, methods=["GET", "POST"]),
    Route(
        "/api/v1/get_job_status_list",
        endpoint=get_job_status_list,
        methods=["GET"],
    ),
    Route("/api/v1/get_tasks_list", endpoint=get_tasks_list, methods=["GET"]),
    Route("/api/v1/get_task_logs", endpoint=get_task_logs, methods=["GET"]),
    Route("/api/v2/validate_csv", endpoint=validate_csv, methods=["POST"]),
    Route(
        "/api/v2/validate_json", endpoint=validate_json_v2, methods=["POST"]
    ),
    Route("/api/v2/submit_jobs", endpoint=submit_jobs_v2, methods=["POST"]),
    Route("/api/v2/cancel_job", endpoint=cancel_job, methods=["POST"]),
    Route("/api/v2/parameters", endpoint=list_parameters_v2, methods=["GET"]),
    Route(
        "/api/v2/parameters/job_types/{job_type:str}/tasks/{task_id:str}",
        endpoint=get_parameter_v2,
        methods=["GET"],
    ),
    Route(
        "/api/v2/parameters/job_types/{job_type:str}/tasks/{task_id:str}"
        "/{modality:str}",
        endpoint=get_parameter_v2,
        methods=["GET"],
    ),
    Route(
        "/api/{version:str}/parameters/job_types/{job_type:str}"
        "/tasks/{task_id:str}",
        endpoint=put_parameter,
        methods=["PUT"],
    ),
    Route(
        "/api/{version:str}/parameters/job_types/{job_type:str}"
        "/tasks/{task_id:str}/{modality:str}",
        endpoint=put_parameter,
        methods=["PUT"],
    ),
    Route("/jobs", endpoint=jobs, methods=["GET"]),
    Route("/job_tasks_table", endpoint=job_tasks_table, methods=["GET"]),
    Route("/task_logs", endpoint=task_logs, methods=["GET"]),
    Route("/job_params", endpoint=job_params, methods=["GET"]),
    Route(
        "/api/job_upload_template",
        endpoint=download_job_template,
        methods=["GET"],
    ),
    Route("/login", login, methods=["GET"]),
    Route("/logout", logout, methods=["GET"]),
    Route("/auth", auth, methods=["GET"]),
    Route("/admin", admin, methods=["GET"]),
]

app = Starlette(routes=routes)
app.add_middleware(SessionMiddleware, secret_key=None)
